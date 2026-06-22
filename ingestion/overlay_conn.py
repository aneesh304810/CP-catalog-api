"""
Business-metadata overlay connector.

Reads STANDARDIZED templates (not free-form docs) so parsing is deterministic:
  - Excel (.xlsx): a 'Datasets' sheet and a 'Columns' sheet with fixed headers
  - Word  (.docx): tables titled 'Datasets' and 'Columns' with the same headers

Produces overlay records keyed by global dataset_key / column_key. These NEVER
overwrite harvested technical metadata; the loader writes them into the
business_desc / owner / sensitivity / tags columns only.

Template contract
-----------------
Datasets sheet/table columns:
    platform_id | schema | object | business_description | owner | tags
Columns sheet/table columns:
    platform_id | schema | object | column | business_description | sensitivity

'platform_id/schema/object[/column]' resolve to the same canonical keys the
technical connectors use, so the overlay lines up exactly.
"""
from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from .model import dataset_key, column_key


@dataclass
class DatasetOverlay:
    dataset_key: str
    business_desc: Optional[str] = None
    owner: Optional[str] = None
    tags: Optional[str] = None


@dataclass
class ColumnOverlay:
    column_key: str
    business_desc: Optional[str] = None
    sensitivity: Optional[str] = None


DATASET_HEADERS = ["platform_id", "schema", "object",
                   "business_description", "owner", "tags"]
COLUMN_HEADERS = ["platform_id", "schema", "object", "column",
                  "business_description", "sensitivity"]


class OverlayConnector:
    def __init__(self, path: str):
        self.path = Path(path)

    def extract(self) -> tuple[list[DatasetOverlay], list[ColumnOverlay]]:
        suffix = self.path.suffix.lower()
        if suffix in (".xlsx", ".xlsm"):
            return self._from_excel()
        if suffix == ".docx":
            return self._from_word()
        raise ValueError(f"Unsupported overlay file type: {suffix}")

    # ---- Excel ---------------------------------------------------------
    def _from_excel(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.path, data_only=True, read_only=True)
        ds_rows = self._sheet_rows(wb, "Datasets")
        col_rows = self._sheet_rows(wb, "Columns")
        return self._build(ds_rows, col_rows)

    def _sheet_rows(self, wb, name) -> list[dict]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower().replace(" ", "_") if h else ""
                   for h in rows[0]]
        out = []
        for r in rows[1:]:
            if not any(c is not None and str(c).strip() for c in r):
                continue
            out.append({headers[i]: (r[i] if i < len(r) else None)
                        for i in range(len(headers))})
        return out

    # ---- Word ----------------------------------------------------------
    def _from_word(self):
        from docx import Document
        doc = Document(self.path)
        ds_rows, col_rows = [], []
        for table in doc.tables:
            rows = [[c.text.strip() for c in row.cells] for row in table.rows]
            if not rows:
                continue
            headers = [h.lower().replace(" ", "_") for h in rows[0]]
            recs = [{headers[i]: (r[i] if i < len(r) else None)
                     for i in range(len(headers))} for r in rows[1:]
                    if any(r)]
            if "column" in headers:
                col_rows.extend(recs)
            elif "object" in headers:
                ds_rows.extend(recs)
        return self._build(ds_rows, col_rows)

    # ---- shared --------------------------------------------------------
    def _build(self, ds_rows, col_rows):
        datasets, columns = [], []
        for r in ds_rows:
            pid, sch, obj = r.get("platform_id"), r.get("schema"), r.get("object")
            if not (pid and obj):
                continue
            datasets.append(DatasetOverlay(
                dataset_key=dataset_key(str(pid), None, sch and str(sch), str(obj)),
                business_desc=_clean(r.get("business_description")),
                owner=_clean(r.get("owner")),
                tags=_clean(r.get("tags")),
            ))
        for r in col_rows:
            pid, sch, obj, col = (r.get("platform_id"), r.get("schema"),
                                  r.get("object"), r.get("column"))
            if not (pid and obj and col):
                continue
            dk = dataset_key(str(pid), None, sch and str(sch), str(obj))
            columns.append(ColumnOverlay(
                column_key=column_key(dk, str(col)),
                business_desc=_clean(r.get("business_description")),
                sensitivity=_clean(r.get("sensitivity")),
            ))
        return datasets, columns


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None
